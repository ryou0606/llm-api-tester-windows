"""Speed test routes with SSE and Excel export."""
import json
from datetime import datetime
from io import BytesIO

from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from starlette.responses import Response

from ..models import SpeedTestRequest
from ..services import data_store, speed_tester

router = APIRouter(prefix="/api", tags=["speed"])


@router.post("/speed-test")
async def run_speed_test(req: SpeedTestRequest):
    """Run speed test and stream results via SSE."""
    configs = await data_store.read_json("api_configs.json")
    if not isinstance(configs, list):
        configs = []

    # Filter to selected configs
    config_map = {c["id"]: c for c in configs}
    selected = []
    for cid in req.config_ids:
        cfg = config_map.get(cid)
        if not cfg:
            continue
        # Override model if specified
        model = req.model_map.get(cid, "")
        if model:
            cfg = {**cfg, "model": model}
        elif cfg.get("models"):
            cfg = {**cfg, "model": cfg["models"][0]}
        else:
            continue
        selected.append(cfg)

    if not selected:
        raise HTTPException(status_code=400, detail="没有可用的配置")

    async def event_stream():
        async for event in speed_tester.run_speed_test(
            configs=selected,
            prompt=req.prompt,
            system_prompt=req.system_prompt,
            rounds=req.rounds,
            concurrency=req.concurrency,
            temperature=req.temperature,
            top_p=req.top_p,
            max_tokens=req.max_tokens,
        ):
            yield event

        # Save to history
        speed_history = await data_store.read_json("speed_history.json")
        if not isinstance(speed_history, list):
            speed_history = []
        # The last event contains the history entry
        # We'll save it from the frontend or handle here

    return StreamingResponse(event_stream(), media_type="text/event-stream",
                             headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


@router.post("/speed-test/save")
async def save_speed_result(entry: dict):
    """Save a speed test result to history."""
    speed_history = await data_store.read_json("speed_history.json")
    if not isinstance(speed_history, list):
        speed_history = []
    speed_history.insert(0, entry)
    # Keep last 200 records
    speed_history = speed_history[:200]
    await data_store.write_json("speed_history.json", speed_history)
    return {"success": True}


@router.get("/export/speed-excel")
async def export_speed_excel():
    """Export speed test history as Excel."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl 未安装")

    speed_history = await data_store.read_json("speed_history.json")
    if not isinstance(speed_history, list):
        speed_history = []

    wb = Workbook()
    ws = wb.active
    ws.title = "测速记录"

    # Header style
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="6C5CE7", end_color="6C5CE7", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="2D3348"),
        right=Side(style="thin", color="2D3348"),
        top=Side(style="thin", color="2D3348"),
        bottom=Side(style="thin", color="2D3348"),
    )

    headers = ["时间", "配置名", "模型", "轮次", "并发", "成功数", "失败数",
               "成功率(%)", "平均TTFB(s)", "TTFB波动", "平均速度(tok/s)",
               "速度波动", "最快耗时(s)", "最慢耗时(s)"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    row = 2
    for entry in speed_history:
        ts = entry.get("timestamp", "")
        for r in entry.get("results", []):
            ws.cell(row=row, column=1, value=ts).border = thin_border
            ws.cell(row=row, column=2, value=r.get("config_name", "")).border = thin_border
            ws.cell(row=row, column=3, value=r.get("model", "")).border = thin_border
            ws.cell(row=row, column=4, value=r.get("total_rounds", 0)).border = thin_border
            ws.cell(row=row, column=5, value=entry.get("params", {}).get("concurrency", 0)).border = thin_border
            ws.cell(row=row, column=6, value=r.get("success", 0)).border = thin_border
            ws.cell(row=row, column=7, value=r.get("fail", 0)).border = thin_border
            ws.cell(row=row, column=8, value=r.get("success_rate", 0)).border = thin_border
            ws.cell(row=row, column=9, value=r.get("avg_ttfb", 0)).border = thin_border
            ws.cell(row=row, column=10, value=r.get("ttfb_stddev", 0)).border = thin_border
            ws.cell(row=row, column=11, value=r.get("avg_speed", 0)).border = thin_border
            ws.cell(row=row, column=12, value=r.get("speed_stddev", 0)).border = thin_border
            ws.cell(row=row, column=13, value=r.get("fastest", 0)).border = thin_border
            ws.cell(row=row, column=14, value=r.get("slowest", 0)).border = thin_border
            row += 1

    # Auto-width columns
    for col_cells in ws.columns:
        max_len = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            try:
                val = str(cell.value or "")
                max_len = max(max_len, len(val.encode("utf-8")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"speed_test_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )
